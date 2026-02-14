#!/usr/bin/env python3
"""
Step 1: Pull live positions, update Excel tracker, write trade plan.

Usage:
    python scripts/ib_update_and_recommend.py

Outputs:
    ~/trading/live_portfolio/trade_plan.csv   (edit this, then run ib_execute_trades.py)
    ~/trading/portfolio_tracking.xlsx         (updated with new performance row)
    ~/trading/live_portfolio/snapshots/       (timestamped backup)
    ~/trading/live_portfolio/trade_plan_archive/  (previous plans)
"""

import csv
import shutil
import sys
from datetime import datetime
from pathlib import Path

import nest_asyncio
import pandas as pd
from ib_insync import IB, Stock
from openpyxl import load_workbook

nest_asyncio.apply()

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))
DATA_DIR = Path.home() / "trade_data" / "ETFTrader"
TRADING_DIR = Path.home() / "trading"
LIVE_DIR = TRADING_DIR / "live_portfolio"
SNAPSHOT_DIR = LIVE_DIR / "snapshots"
ARCHIVE_DIR = LIVE_DIR / "trade_plan_archive"
TRADE_PLAN_FILE = LIVE_DIR / "trade_plan.csv"
EXCEL_FILE = TRADING_DIR / "portfolio_tracking.xlsx"

IB_HOST = "127.0.0.1"
IB_PORT = 4001
IB_CLIENT_ID = 1

ENTRY_STOP_LOSS_PCT = 0.12
CASH_RESERVE = 70_000  # Minimum cash reserve - trades must not breach this


# ── IB helpers (reused from ib_portfolio_manager.py) ──────────


def connect_ib():
    ib = IB()
    print(f"Connecting to IB Gateway ({IB_HOST}:{IB_PORT})...")
    try:
        ib.connect(
            IB_HOST, IB_PORT,
            clientId=IB_CLIENT_ID, readonly=True, timeout=10,
        )
    except Exception as e:
        print(f"FAILED: {e}")
        sys.exit(1)
    accts = ib.managedAccounts()
    print(f"Connected. Account: {accts[0]}\n")
    return ib


def get_account_summary(ib):
    summary = {}
    for av in ib.accountSummary():
        if av.currency == "USD":
            try:
                summary[av.tag] = float(av.value)
            except (ValueError, TypeError):
                pass
    return summary


def get_positions(ib):
    return [
        {
            "ticker": p.contract.symbol,
            "shares": float(p.position),
            "avg_cost": float(p.avgCost),
        }
        for p in ib.positions()
    ]


def get_live_prices(ib, tickers):
    prices = {}
    qualified = []
    for t in tickers:
        c = Stock(t, "SMART", "USD")
        try:
            ib.qualifyContracts(c)
            if c.conId:
                qualified.append(c)
        except Exception:
            pass

    if not qualified:
        return prices

    ib.reqMarketDataType(3)  # delayed = free
    snaps = [(c.symbol, ib.reqMktData(c, snapshot=True)) for c in qualified]
    ib.sleep(4)

    for sym, td in snaps:
        for attr in ("last", "close", "bid", "ask"):
            v = getattr(td, attr, None)
            if v is not None and v == v and v > 0:
                prices[sym] = float(v)
                break
        ib.cancelMktData(td.contract)

    ib.reqMarketDataType(1)
    return prices


def load_target_portfolio():
    path = LIVE_DIR / "target_portfolio_latest.csv"
    if not path.exists():
        return {}
    df = pd.read_csv(path, index_col=0)
    return {t: float(r.iloc[0]) for t, r in df.iterrows()}


def load_factor_scores():
    path = DATA_DIR / "factor_scores_latest.parquet"
    if not path.exists():
        return pd.Series(dtype=float)
    df = pd.read_parquet(path)
    return df.iloc[:, 0].sort_values(ascending=False)


# ── Build position data ──────────────────────────────────────


def build_pos_data(ib_positions, live_prices, targets, scores, nlv):
    pos_data = []
    for pos in ib_positions:
        t = pos["ticker"]
        if pos["shares"] == 0:
            continue
        price = live_prices.get(t, pos["avg_cost"])
        mv = pos["shares"] * price
        cb = pos["shares"] * pos["avg_cost"]
        pnl = mv - cb
        pnl_pct = pnl / cb if cb != 0 else 0
        tw = targets.get(t, 0)
        aw = mv / nlv if nlv else 0
        score = scores.get(t) if t in scores.index else None
        pos_data.append({
            "ticker": t,
            "shares": pos["shares"],
            "avg_cost": pos["avg_cost"],
            "price": price,
            "mkt_value": mv,
            "pnl": pnl,
            "pnl_pct": pnl_pct,
            "target_w": tw,
            "actual_w": aw,
            "drift": aw - tw,
            "score": score,
            "stop": pos["avg_cost"] * (1 - ENTRY_STOP_LOSS_PCT),
            "in_target": t in targets,
        })
    pos_data.sort(key=lambda x: (-x["target_w"], -x["mkt_value"]))
    return pos_data


# ── Generate trade recommendations ───────────────────────────


def make_trades(pos_data, targets, live_prices, nlv, cash):
    ib_map = {p["ticker"]: p for p in pos_data}
    trades = []
    special = {"IBKR"}

    # Sell non-target positions
    for p in pos_data:
        if not p["in_target"] and p["ticker"] not in special:
            if p["shares"] > 0:
                trades.append({
                    "action": "SELL",
                    "ticker": p["ticker"],
                    "shares": int(abs(p["shares"])),
                    "price": round(p["price"], 2),
                    "est_value": round(abs(p["mkt_value"]), 2),
                    "reason": "Not in strategy target",
                })
            elif p["shares"] < 0:
                trades.append({
                    "action": "BUY_TO_COVER",
                    "ticker": p["ticker"],
                    "shares": int(abs(p["shares"])),
                    "price": round(p["price"], 2),
                    "est_value": round(abs(p["mkt_value"]), 2),
                    "reason": "Close short position",
                })

    # Calculate available cash for buys (cash - reserve + expected sell proceeds)
    sell_proceeds = sum(
        t["est_value"] for t in trades if t["action"] == "SELL"
    )
    # BUY_TO_COVER costs money (closing shorts)
    cover_cost = sum(
        t["est_value"] for t in trades if t["action"] == "BUY_TO_COVER"
    )
    available_for_buys = cash + sell_proceeds - cover_cost - CASH_RESERVE

    if available_for_buys < 0:
        available_for_buys = 0

    # Collect all potential buy trades first, then cap by available cash
    buy_candidates = []

    # Buy missing targets
    for ticker, tw in targets.items():
        if ticker not in ib_map:
            price = live_prices.get(ticker)
            if price and price > 0:
                tgt_val = tw * nlv
                sh = int(tgt_val / price)
                if sh > 0:
                    buy_candidates.append({
                        "action": "BUY",
                        "ticker": ticker,
                        "shares": sh,
                        "price": round(price, 2),
                        "est_value": round(sh * price, 2),
                        "reason": f"New position (target {tw*100:.0f}%)",
                    })

    # Rebalance drifted positions (>3%)
    for p in pos_data:
        if p["target_w"] > 0 and abs(p["drift"]) > 0.03:
            diff = p["target_w"] * nlv - p["mkt_value"]
            ds = int(abs(diff) / p["price"])
            if ds > 0:
                if diff > 0:
                    buy_candidates.append({
                        "action": "BUY",
                        "ticker": p["ticker"],
                        "shares": ds,
                        "price": round(p["price"], 2),
                        "est_value": round(ds * p["price"], 2),
                        "reason": f"Rebalance (drift {p['drift']*100:+.1f}%)",
                    })
                else:
                    trades.append({
                        "action": "SELL",
                        "ticker": p["ticker"],
                        "shares": ds,
                        "price": round(p["price"], 2),
                        "est_value": round(ds * p["price"], 2),
                        "reason": f"Rebalance (drift {p['drift']*100:+.1f}%)",
                    })

    # Apply cash reserve cap to buy candidates
    running_spend = 0
    for bc in buy_candidates:
        if running_spend + bc["est_value"] <= available_for_buys:
            trades.append(bc)
            running_spend += bc["est_value"]
        else:
            # Reduce shares to fit within available cash
            remaining = available_for_buys - running_spend
            if remaining > bc["price"]:
                reduced_shares = int(remaining / bc["price"])
                if reduced_shares > 0:
                    bc["shares"] = reduced_shares
                    bc["est_value"] = round(reduced_shares * bc["price"], 2)
                    bc["reason"] += f" [capped by ${CASH_RESERVE:,.0f} reserve]"
                    trades.append(bc)
                    running_spend += bc["est_value"]
            # If nothing fits, skip this trade entirely

    trades.sort(
        key=lambda t: (0 if "SELL" in t["action"] else 1, t["ticker"])
    )
    return trades


# ── Write trade plan file ─────────────────────────────────────


def archive_trade_plan():
    """Archive existing trade plan before writing a new one."""
    if not TRADE_PLAN_FILE.exists():
        return
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    mtime = datetime.fromtimestamp(TRADE_PLAN_FILE.stat().st_mtime)
    archive_name = f"trade_plan_{mtime.strftime('%Y%m%d_%H%M%S')}.csv"
    dest = ARCHIVE_DIR / archive_name
    shutil.copy2(TRADE_PLAN_FILE, dest)
    print(f"  Archived previous plan: {archive_name}")


def write_trade_plan(trades):
    LIVE_DIR.mkdir(parents=True, exist_ok=True)
    archive_trade_plan()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    with open(TRADE_PLAN_FILE, "w", newline="") as f:
        f.write(f"# TRADE PLAN - Generated {ts}\n")
        f.write("#\n")
        f.write("# HOW TO USE:\n")
        f.write("#   1. Review each trade below\n")
        f.write("#   2. Edit the 'instruction' column:\n")
        f.write("#        APPROVE  = execute as-is\n")
        f.write("#        SKIP     = do not execute\n")
        f.write("#        Or write plain English, e.g.:\n")
        f.write("#          reduce to 30 shares\n")
        f.write("#          change to limit order at $85\n")
        f.write("#          buy 100 shares instead\n")
        f.write("#          only if price drops below $40\n")
        f.write("#   3. Save this file\n")
        f.write(
            "#   4. Run: python scripts/ib_execute_trades.py\n"
        )
        f.write("#\n")

        writer = csv.DictWriter(
            f,
            fieldnames=[
                "action", "ticker", "shares", "price",
                "est_value", "reason", "instruction",
            ],
        )
        writer.writeheader()
        for t in trades:
            t["instruction"] = "APPROVE"
            writer.writerow(t)

    print(f"\n  Trade plan written: {TRADE_PLAN_FILE}")
    print(f"  {len(trades)} trade(s) ready for review.")
    print(f"\n  Next steps:")
    print(f"    1. Edit {TRADE_PLAN_FILE.name}")
    print(f"    2. Run: python scripts/ib_execute_trades.py")


# ── Update Excel tracker ──────────────────────────────────────


def update_excel(pos_data, summary):
    if not EXCEL_FILE.exists():
        print(f"  Excel file not found: {EXCEL_FILE}")
        print("  Skipping Excel update.")
        return

    ts = datetime.now()
    nlv = summary.get("NetLiquidation", 0)
    upnl = summary.get("UnrealizedPnL", 0)
    n_pos = len([p for p in pos_data if p["shares"] != 0])

    try:
        wb = load_workbook(EXCEL_FILE)

        # ── Update Performance sheet ──
        if "Performance" in wb.sheetnames:
            ws = wb["Performance"]
            # Find next empty row
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=ts)
            ws.cell(row=row, column=2, value=nlv)
            ws.cell(row=row, column=3, value=n_pos)
            # expected_return, expected_volatility, expected_sharpe
            # leave blank or carry forward
            print(f"  Performance sheet: added row {row} "
                  f"(NLV=${nlv:,.0f}, {n_pos} positions)")

        # ── Update Positions sheet with current snapshot ──
        if "Positions" in wb.sheetnames:
            ws = wb["Positions"]
            for p in pos_data:
                if p["shares"] == 0:
                    continue
                row = ws.max_row + 1
                ws.cell(row=row, column=1, value=ts)
                ws.cell(row=row, column=2, value=p["ticker"])
                ws.cell(row=row, column=3, value="")  # name
                ws.cell(row=row, column=4, value=round(p["actual_w"], 4))
                ws.cell(row=row, column=5, value=round(p["mkt_value"], 2))
                ws.cell(row=row, column=6, value=p["shares"])
                ws.cell(row=row, column=7, value=round(p["price"], 2))
                ws.cell(
                    row=row, column=8,
                    value=round(p["score"], 6) if p["score"] else None
                )

            print(f"  Positions sheet: added {n_pos} rows")

        wb.save(EXCEL_FILE)
        print(f"  Saved: {EXCEL_FILE.name}")

    except Exception as e:
        print(f"  Error updating Excel: {e}")


# ── Save snapshot ─────────────────────────────────────────────


def save_snapshot(pos_data, summary):
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    pos_file = SNAPSHOT_DIR / f"positions_{ts}.csv"
    fields = [
        "ticker", "shares", "avg_cost", "price",
        "mkt_value", "pnl", "pnl_pct", "target_w",
        "actual_w", "drift", "score", "stop",
    ]
    with open(pos_file, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for p in pos_data:
            w.writerow({
                k: (round(p[k], 4)
                    if isinstance(p[k], float) and p[k] is not None
                    else p[k])
                for k in fields
                if k in p
            })

    acct_file = SNAPSHOT_DIR / f"account_{ts}.csv"
    with open(acct_file, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["metric", "value"])
        for k, v in summary.items():
            w.writerow([k, v])

    latest = LIVE_DIR / "positions_latest.csv"
    shutil.copy2(pos_file, latest)
    print(f"  Snapshot saved: {pos_file.name}")


# ── Display ───────────────────────────────────────────────────


def display_summary(pos_data, summary, trades):
    nlv = summary.get("NetLiquidation", 0)
    cash = summary.get("TotalCashValue", 0)
    posv = summary.get("GrossPositionValue", 0)

    print("\n" + "=" * 70)
    print(f"  PORTFOLIO SNAPSHOT  |  "
          f"{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 70)
    print(f"  NLV: ${nlv:>12,.2f}   Cash: ${cash:>12,.2f}   "
          f"Positions: ${posv:>12,.2f}")
    invested_pct = posv / nlv * 100 if nlv else 0
    print(f"  Invested: {invested_pct:.1f}%")

    # Cash reserve status
    reserve_ok = cash >= CASH_RESERVE
    reserve_status = "OK" if reserve_ok else "BELOW MINIMUM"
    print(f"\n  Cash Reserve:  ${CASH_RESERVE:>12,.0f} minimum")
    print(f"  Current Cash:  ${cash:>12,.2f}  [{reserve_status}]")
    if not reserve_ok:
        shortfall = CASH_RESERVE - cash
        print(f"  SHORTFALL:     ${shortfall:>12,.2f}  "
              f"** Buy trades will be restricted **")

    # Quick position summary
    strategy = [p for p in pos_data if p["in_target"]]
    total_pnl = sum(p["pnl"] for p in pos_data)
    total_cost = sum(abs(p["shares"] * p["avg_cost"]) for p in pos_data)
    ret = total_pnl / total_cost if total_cost else 0

    print(f"\n  Positions: {len(pos_data)} total, "
          f"{len(strategy)} strategy")
    winners = sum(1 for p in pos_data if p["pnl"] > 0)
    print(f"  Winners: {winners}/{len(pos_data)}  "
          f"Total P&L: ${total_pnl:>+,.0f} ({ret:+.1%})")

    # Quick trade summary
    if trades:
        buys = sum(
            t["est_value"] for t in trades if "BUY" in t["action"]
        )
        sells = sum(
            t["est_value"] for t in trades
            if t["action"] == "SELL"
        )
        net_cash_after = cash + sells - buys
        print(f"\n  Trade plan: {len(trades)} trades  "
              f"(buys ${buys:,.0f} / sells ${sells:,.0f})")
        print(f"  Cash after trades: ${net_cash_after:>,.0f}  "
              f"(reserve ${CASH_RESERVE:,.0f})")
    else:
        print("\n  No trades recommended.")


# ── Main ──────────────────────────────────────────────────────


def main():
    ib = connect_ib()

    try:
        summary = get_account_summary(ib)
        ib_positions = get_positions(ib)
        nlv = summary.get("NetLiquidation", 0)

        targets = load_target_portfolio()
        scores = load_factor_scores()

        all_tickers = sorted(
            set(p["ticker"] for p in ib_positions) | set(targets.keys())
        )
        print(f"Fetching prices for {len(all_tickers)} tickers...")
        prices = get_live_prices(ib, all_tickers)
        print(f"Got {len(prices)} prices.\n")

        cash = summary.get("TotalCashValue", 0)

        pos_data = build_pos_data(
            ib_positions, prices, targets, scores, nlv
        )
        trades = make_trades(pos_data, targets, prices, nlv, cash)

        display_summary(pos_data, summary, trades)

        print("\n--- UPDATING RECORDS " + "-" * 49)
        save_snapshot(pos_data, summary)
        update_excel(pos_data, summary)
        write_trade_plan(trades)

    finally:
        ib.disconnect()
        print("\nDisconnected from IB Gateway.")


if __name__ == "__main__":
    main()
