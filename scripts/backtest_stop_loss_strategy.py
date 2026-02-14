#!/usr/bin/env python3
"""
Stop-Loss Momentum Strategy Backtest

Strategy:
- Select top 20 ETFs by factor scores bi-weekly
- Hold positions as long as they're above entry price
- Sell immediately if price drops below entry (stop-loss)
- Use trailing stops for big winners (>10% gain)
- "Let winners run, cut losers quickly"

Based on academic research: Han, Zhou, Zhu (2014) - "Taming Momentum Crashes"

Usage:
    python scripts/backtest_stop_loss_strategy.py --start 2025-03-01 --capital 100000 --monthly-add 5000
"""

import argparse
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path
from dataclasses import dataclass, field
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.factors import (
    FactorIntegrator,
    MomentumFactor,
    QualityFactor,
    SimplifiedValueFactor,
    VolatilityFactor,
)
from src.portfolio import MeanVarianceOptimizer

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# Constants
DATA_DIR = PROJECT_ROOT / "data"
RESULTS_DIR = Path.home() / "trading"


@dataclass
class Position:
    """Track individual position with stop-loss state"""
    ticker: str
    shares: float
    entry_price: float
    entry_date: pd.Timestamp
    peak_price: float = field(init=False)
    peak_date: pd.Timestamp = field(init=False)
    flagged: bool = False

    def __post_init__(self):
        self.peak_price = self.entry_price
        self.peak_date = self.entry_date

    def update_peak(self, current_price: float, current_date: pd.Timestamp):
        """Update peak price if current price is higher"""
        if current_price > self.peak_price:
            self.peak_price = current_price
            self.peak_date = current_date

    def current_gain(self, current_price: float) -> float:
        """Calculate gain from entry"""
        return (current_price - self.entry_price) / self.entry_price

    def drawdown_from_peak(self, current_price: float) -> float:
        """Calculate drawdown from peak"""
        return (current_price - self.peak_price) / self.peak_price

    def days_held(self, current_date: pd.Timestamp) -> int:
        """Calculate days held"""
        return (current_date - self.entry_date).days


class StopLossBacktest:
    """Backtest with stop-loss based exits and bi-weekly factor updates"""

    def __init__(
        self,
        start_date: str,
        end_date: str,
        initial_capital: float,
        monthly_addition: float,
        num_positions: int = 20,
        # Capital deployment parameters
        immediate_deploy_pct: float = 0.30,     # Deploy 30% immediately
        monthly_deploy_from_reserve: float = 10000,  # Deploy $10k/month from reserve
        fixed_income_ticker: str = "SGOV",      # ETF for cash reserve
        # Stop-loss parameters
        use_entry_price_stop: bool = True,
        use_trailing_stop: bool = False,          # NEW: Pure trailing stop from peak
        trailing_stop_pct: float = 0.10,          # NEW: Trail distance (e.g., 0.10 = -10% from peak)
        trailing_stop_threshold: float = 0.10,  # Activate at +10% gain
        trailing_stop_distance: float = 0.08,   # Trail by 8%
        factor_recalc_weeks: int = 2,           # Recalc factors every 2 weeks
    ):
        self.start_date = pd.Timestamp(start_date)
        self.end_date = pd.Timestamp(end_date)
        self.initial_capital = initial_capital
        self.monthly_addition = monthly_addition
        self.num_positions = num_positions

        # Capital deployment parameters
        self.immediate_deploy_pct = immediate_deploy_pct
        self.monthly_deploy_from_reserve = monthly_deploy_from_reserve
        self.fixed_income_ticker = fixed_income_ticker

        # Stop-loss parameters
        self.use_entry_price_stop = use_entry_price_stop
        self.use_trailing_stop = use_trailing_stop
        self.trailing_stop_pct = trailing_stop_pct
        self.entry_stop_loss_pct = 0.12  # -12% stop from entry price
        self.trailing_stop_threshold = trailing_stop_threshold
        self.trailing_stop_distance = trailing_stop_distance
        self.factor_recalc_weeks = factor_recalc_weeks

        # State
        self.positions: Dict[str, Position] = {}
        self.fixed_income_position: Optional[Position] = None  # Track SGOV/BIL holding
        self.cash = 0  # Start with zero cash, all goes into immediate deploy or fixed income
        self.total_contributions = initial_capital

        # History tracking
        self.portfolio_history = []
        self.trades_history = []
        self.performance_history = []
        self.stop_loss_history = []

        # Caches
        self._etf_names_cache = {}

        # Initialize factors
        self.factor_calculators = {
            "momentum": MomentumFactor(lookback=252, skip_recent=21),
            "quality": QualityFactor(lookback=252),
            "value": SimplifiedValueFactor(),
            "volatility": VolatilityFactor(lookback=60),
        }
        self.integrator = FactorIntegrator(
            factor_weights={
                "momentum": 0.25,
                "quality": 0.25,
                "value": 0.25,
                "volatility": 0.25,
            }
        )

        logger.info(f"Stop-Loss Strategy initialized:")
        logger.info(f"  Capital deployment: {immediate_deploy_pct:.0%} immediate, ${monthly_deploy_from_reserve:,.0f}/month from reserve")
        logger.info(f"  Fixed income ticker: {fixed_income_ticker}")
        logger.info(f"  Entry stop-loss: {use_entry_price_stop} (-{self.entry_stop_loss_pct:.0%} from entry)")
        logger.info(f"  Pure trailing stop: {use_trailing_stop} (-{trailing_stop_pct:.0%} from peak)")
        logger.info(f"  Gain-triggered trailing: {trailing_stop_threshold:.0%} gain, {trailing_stop_distance:.0%} trail")
        logger.info(f"  Factor recalc: Every {factor_recalc_weeks} weeks")

    def _get_etf_name(self, ticker: str) -> str:
        """Get ETF name with caching"""
        if ticker in self._etf_names_cache:
            return self._etf_names_cache[ticker]

        try:
            etf = yf.Ticker(ticker)
            name = etf.info.get('longName') or etf.info.get('shortName') or ticker
            self._etf_names_cache[ticker] = name
            return name
        except:
            self._etf_names_cache[ticker] = ticker
            return ticker

    def _calculate_factors(self, prices: pd.DataFrame, date: pd.Timestamp) -> pd.Series:
        """Calculate factor scores at given date"""
        historical_prices = prices.loc[:date]

        if len(historical_prices) < 252:
            return pd.Series(dtype=float)

        factor_scores_dict = {}
        for factor_name, calculator in self.factor_calculators.items():
            try:
                if factor_name == "value":
                    expense_ratios = pd.Series(
                        np.random.uniform(0.0005, 0.01, len(historical_prices.columns)),
                        index=historical_prices.columns,
                    )
                    scores = calculator.calculate(historical_prices, expense_ratios)
                else:
                    scores = calculator.calculate(historical_prices)
                factor_scores_dict[factor_name] = scores
            except:
                factor_scores_dict[factor_name] = pd.Series(dtype=float)

        factor_scores_df = pd.DataFrame(factor_scores_dict)
        integrated_scores = self.integrator.integrate(factor_scores_df)

        return integrated_scores

    def _check_stop_losses(self, prices: pd.DataFrame, date: pd.Timestamp) -> List[str]:
        """Check all positions for stop-loss triggers, return tickers to sell"""
        to_sell = []
        current_prices = prices.loc[date]

        for ticker, position in self.positions.items():
            if ticker not in current_prices.index or pd.isna(current_prices[ticker]):
                continue

            current_price = current_prices[ticker]

            # Update peak
            position.update_peak(current_price, date)

            # Check pure trailing stop (from peak, no gain threshold required)
            if self.use_trailing_stop:
                drawdown_from_peak = position.drawdown_from_peak(current_price)
                if drawdown_from_peak < -self.trailing_stop_pct:
                    to_sell.append((ticker, f"Trail stop (-{self.trailing_stop_pct:.0%} from peak)", current_price))
                    logger.info(f"  📉 TRAIL STOP: {ticker} at ${current_price:.2f} (peak: ${position.peak_price:.2f}, dd: {drawdown_from_peak:.1%})")
                    continue

            # Check entry price stop-loss (-12% from entry)
            loss_pct = position.current_gain(current_price)
            if self.use_entry_price_stop and loss_pct < -self.entry_stop_loss_pct:
                to_sell.append((ticker, f"Stop-loss (-{self.entry_stop_loss_pct:.0%})", current_price))
                logger.info(f"  🛑 STOP-LOSS: {ticker} at ${current_price:.2f} (entry: ${position.entry_price:.2f}, loss: {loss_pct:.1%})")
                continue

            # Check trailing stop for winners
            gain = position.current_gain(current_price)
            if gain > self.trailing_stop_threshold:
                drawdown = position.drawdown_from_peak(current_price)
                if drawdown < -self.trailing_stop_distance:
                    to_sell.append((ticker, f"Trailing stop (peak: ${position.peak_price:.2f})", current_price))
                    logger.info(f"  📉 TRAILING STOP: {ticker} at ${current_price:.2f} (peak: ${position.peak_price:.2f}, gain: {gain:.1%}, dd from peak: {drawdown:.1%})")

        return to_sell

    def _execute_stop_loss_sells(self, to_sell: List, date: pd.Timestamp):
        """Execute stop-loss sells"""
        for ticker, reason, price in to_sell:
            position = self.positions[ticker]
            value = position.shares * price
            self.cash += value

            # Record trade
            self.trades_history.append({
                'date': date,
                'ticker': ticker,
                'action': 'SELL_STOP',
                'shares': position.shares,
                'price': price,
                'value': value,
                'reason': reason,
                'gain': position.current_gain(price),
                'days_held': position.days_held(date),
            })

            # Record stop-loss event
            self.stop_loss_history.append({
                'date': date,
                'ticker': ticker,
                'entry_price': position.entry_price,
                'exit_price': price,
                'peak_price': position.peak_price,
                'gain': position.current_gain(price),
                'reason': reason,
            })

            # Remove position
            del self.positions[ticker]

    def _fill_empty_slots(self, factor_scores: pd.Series, prices: pd.DataFrame, date: pd.Timestamp):
        """Buy new positions to fill empty slots"""
        current_tickers = set(self.positions.keys())
        num_empty_slots = self.num_positions - len(self.positions)

        if num_empty_slots <= 0:
            return

        # Get available capital per position
        total_value = self.get_portfolio_value(prices, date) + self.cash
        target_position_value = total_value / self.num_positions

        # Rank candidates (exclude current positions)
        valid_scores = factor_scores.dropna()
        candidates = valid_scores[~valid_scores.index.isin(current_tickers)]
        candidates = candidates.sort_values(ascending=False)

        # Buy top candidates
        current_prices = prices.loc[date]
        bought = 0

        for ticker in candidates.index[:num_empty_slots * 2]:  # Check 2x to handle failures
            if bought >= num_empty_slots:
                break

            if ticker not in current_prices.index or pd.isna(current_prices[ticker]):
                continue

            price = current_prices[ticker]
            if price <= 0 or self.cash < target_position_value * 0.5:  # Need at least half position value
                continue

            # Calculate shares
            shares = int(target_position_value / price)
            if shares == 0:
                continue

            cost = shares * price
            if cost > self.cash:
                shares = int(self.cash / price)
                cost = shares * price

            if shares > 0:
                # Create position
                self.positions[ticker] = Position(
                    ticker=ticker,
                    shares=shares,
                    entry_price=price,
                    entry_date=date,
                )
                self.cash -= cost
                bought += 1

                # Record trade
                self.trades_history.append({
                    'date': date,
                    'ticker': ticker,
                    'action': 'BUY',
                    'shares': shares,
                    'price': price,
                    'value': cost,
                    'reason': 'Fill slot',
                    'factor_score': factor_scores[ticker],
                })

                logger.info(f"  ✅ BUY: {ticker} x {shares} @ ${price:.2f} = ${cost:,.2f} (score: {factor_scores[ticker]:.3f})")

    def get_portfolio_value(self, prices: pd.DataFrame, date: pd.Timestamp) -> float:
        """Calculate current portfolio value"""
        if len(self.positions) == 0:
            return 0.0

        current_prices = prices.loc[date]
        total = 0.0

        for ticker, position in self.positions.items():
            if ticker in current_prices.index and not pd.isna(current_prices[ticker]):
                total += position.shares * current_prices[ticker]

        return total

    def run(self):
        """Run complete backtest"""
        logger.info("=" * 80)
        logger.info("STOP-LOSS STRATEGY BACKTEST")
        logger.info("=" * 80)

        # Load prices
        prices_file = DATA_DIR / "processed" / "etf_prices_filtered.parquet"
        if not prices_file.exists():
            logger.error(f"Price data not found: {prices_file}")
            return False

        prices = pd.read_parquet(prices_file)
        logger.info(f"Loaded {len(prices.columns)} ETFs, {len(prices)} days")

        # Get weekly dates
        start_date = self.start_date
        end_date = self.end_date

        if prices.index.tz is not None:
            if start_date.tz is None:
                start_date = start_date.tz_localize(prices.index.tz)
            if end_date.tz is None:
                end_date = end_date.tz_localize(prices.index.tz)

        trading_days = prices.index[(prices.index >= start_date) & (prices.index <= end_date)]

        # Get weekly check dates (first day of each week)
        check_dates = []
        current_week = None
        for date in trading_days:
            week_key = (date.isocalendar()[0], date.isocalendar()[1])
            if week_key != current_week:
                check_dates.append(date)
                current_week = week_key

        logger.info(f"Running {len(check_dates)} weekly checks from {check_dates[0].date()} to {check_dates[-1].date()}")

        # Initialize capital deployment
        immediate_deploy = self.initial_capital * self.immediate_deploy_pct
        reserve = self.initial_capital - immediate_deploy
        self.cash = immediate_deploy  # Start with immediate deployment amount
        reserve_remaining = reserve  # Track reserve for monthly deployment

        logger.info(f"💰 Initial capital deployment:")
        logger.info(f"  Immediate: ${immediate_deploy:,.2f} ({self.immediate_deploy_pct:.1%})")
        logger.info(f"  Reserve ({self.fixed_income_ticker}): ${reserve:,.2f} ({1-self.immediate_deploy_pct:.1%})")

        # Run backtest
        last_factor_calc_week = -999
        last_deployment_month = None

        for i, date in enumerate(check_dates):
            week_num = date.isocalendar()[1]
            logger.info("=" * 80)
            logger.info(f"Week {i+1}/{len(check_dates)}: {date.date()}")
            logger.info("=" * 80)

            # Deploy from reserve (monthly)
            current_month = (date.year, date.month)
            if i > 0 and current_month != last_deployment_month and reserve_remaining > 0:
                deploy_amount = min(self.monthly_deploy_from_reserve, reserve_remaining)
                self.cash += deploy_amount
                reserve_remaining -= deploy_amount
                last_deployment_month = current_month
                logger.info(f"💰 Deployed ${deploy_amount:,.2f} from reserve (remaining: ${reserve_remaining:,.2f})")

            # Add monthly capital contribution (if applicable)
            if i > 0 and self._is_last_week_of_month(date) and self.monthly_addition > 0:
                self.cash += self.monthly_addition
                self.total_contributions += self.monthly_addition
                logger.info(f"💰 Added ${self.monthly_addition:,.2f} monthly contribution")

            # Check stop-losses (every week)
            to_sell = self._check_stop_losses(prices, date)
            if len(to_sell) > 0:
                logger.info(f"Stop-losses triggered: {len(to_sell)}")
                self._execute_stop_loss_sells(to_sell, date)

            # Recalculate factors (every N weeks)
            weeks_since_last = week_num - last_factor_calc_week
            should_recalc = (last_factor_calc_week == -999) or (weeks_since_last >= self.factor_recalc_weeks)

            if should_recalc:
                logger.info(f"🔄 Recalculating factors (every {self.factor_recalc_weeks} weeks)")
                factor_scores = self._calculate_factors(prices, date)
                last_factor_calc_week = week_num

                # Fill empty slots
                if len(self.positions) < self.num_positions:
                    logger.info(f"Filling {self.num_positions - len(self.positions)} empty slots")
                    self._fill_empty_slots(factor_scores, prices, date)

            # Calculate performance
            portfolio_value = self.get_portfolio_value(prices, date)
            total_value = portfolio_value + self.cash

            logger.info(f"Portfolio: ${portfolio_value:,.2f} | Cash: ${self.cash:,.2f} | Total: ${total_value:,.2f}")
            logger.info(f"Positions: {len(self.positions)}/{self.num_positions}")

            # Track performance
            if i == 0:
                period_return = 0.0
                cumulative_return = 0.0
            else:
                prev_data = self.performance_history[-1]
                prev_value = prev_data['total_value']
                capital_added = self.monthly_addition if self._is_last_week_of_month(date) else 0
                period_return = (total_value - prev_value - capital_added) / prev_value if prev_value > 0 else 0
                cumulative_return = (total_value - self.total_contributions) / self.total_contributions

            self.performance_history.append({
                'date': date,
                'portfolio_value': portfolio_value,
                'cash': self.cash,
                'total_value': total_value,
                'contributions': self.total_contributions,
                'period_return': period_return,
                'cumulative_return': cumulative_return,
                'num_positions': len(self.positions),
            })

            logger.info(f"Period return: {period_return:+.2%} | Cumulative: {cumulative_return:+.2%}")

            # Save portfolio snapshot
            if len(self.positions) > 0:
                positions_data = []
                current_prices = prices.loc[date]
                for ticker, pos in self.positions.items():
                    if ticker in current_prices.index:
                        current_price = current_prices[ticker]
                        positions_data.append({
                            'date': date,
                            'ticker': ticker,
                            'shares': pos.shares,
                            'entry_price': pos.entry_price,
                            'current_price': current_price,
                            'value': pos.shares * current_price,
                            'gain': pos.current_gain(current_price),
                            'days_held': pos.days_held(date),
                            'peak_price': pos.peak_price,
                        })
                self.portfolio_history.extend(positions_data)

        logger.info("=" * 80)
        logger.info("BACKTEST COMPLETE")
        logger.info("=" * 80)

        return True

    def _is_last_week_of_month(self, date: pd.Timestamp) -> bool:
        """Check if date is in last week of month"""
        if date.month == 12:
            next_month = date.replace(year=date.year + 1, month=1, day=1)
        else:
            next_month = date.replace(month=date.month + 1, day=1)
        last_day = next_month - timedelta(days=1)
        return (last_day - date).days < 7

    def generate_results(self) -> dict:
        """Calculate final metrics"""
        perf_df = pd.DataFrame(self.performance_history)

        returns = perf_df['period_return'].values[1:]
        final_value = perf_df['total_value'].iloc[-1]
        total_contrib = perf_df['contributions'].iloc[-1]

        days = (perf_df['date'].iloc[-1] - perf_df['date'].iloc[0]).days
        years = days / 365.25
        total_return = (final_value - total_contrib) / total_contrib
        cagr = (1 + total_return) ** (1 / years) - 1 if years > 0 else 0

        volatility = np.std(returns) * np.sqrt(52)
        mean_return = np.mean(returns) * 52
        sharpe = mean_return / volatility if volatility > 0 else 0

        cumulative_values = perf_df['total_value'].values
        peak = np.maximum.accumulate(cumulative_values)
        drawdown = (cumulative_values - peak) / peak
        max_drawdown = np.min(drawdown)

        # Stop-loss stats
        stop_loss_df = pd.DataFrame(self.stop_loss_history)
        if len(stop_loss_df) > 0:
            avg_loss = stop_loss_df['gain'].mean()
            num_stops = len(stop_loss_df)
        else:
            avg_loss = 0
            num_stops = 0

        # Trade stats
        trades_df = pd.DataFrame(self.trades_history)
        if len(trades_df) > 0:
            sell_trades = trades_df[trades_df['action'] == 'SELL_STOP']
            if len(sell_trades) > 0:
                winners = (sell_trades['gain'] > 0).sum()
                losers = (sell_trades['gain'] <= 0).sum()
                win_rate = winners / (winners + losers) if (winners + losers) > 0 else 0
            else:
                win_rate = 0
        else:
            win_rate = 0

        return {
            'start_date': perf_df['date'].iloc[0],
            'end_date': perf_df['date'].iloc[-1],
            'initial_capital': self.initial_capital,
            'total_contributions': total_contrib,
            'final_value': final_value,
            'total_return': total_return,
            'cagr': cagr,
            'volatility': volatility,
            'sharpe_ratio': sharpe,
            'max_drawdown': max_drawdown,
            'num_stop_losses': num_stops,
            'avg_stop_loss': avg_loss,
            'win_rate': win_rate,
            'total_trades': len(trades_df),
        }

    def save_to_excel(self, filename: str):
        """Save results to Excel with detailed position tracking"""
        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Summary
        ws = wb.create_sheet("Summary")
        results = self.generate_results()

        for idx, (key, value) in enumerate(results.items(), 1):
            ws.cell(row=idx, column=1, value=key)
            if isinstance(value, (int, float)):
                ws.cell(row=idx, column=2, value=value)
            else:
                ws.cell(row=idx, column=2, value=str(value))

        # Performance
        ws = wb.create_sheet("Performance")
        perf_df = pd.DataFrame(self.performance_history)
        if 'date' in perf_df.columns:
            perf_df['date'] = pd.to_datetime(perf_df['date']).dt.tz_localize(None)

        for r_idx, row in enumerate(dataframe_to_rows(perf_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Positions - NEW: Weekly position snapshots with ETF names
        if len(self.portfolio_history) > 0:
            ws = wb.create_sheet("Positions")
            positions_df = pd.DataFrame(self.portfolio_history)
            if 'date' in positions_df.columns:
                positions_df['date'] = pd.to_datetime(positions_df['date']).dt.tz_localize(None)

            # Add ETF names column
            logger.info("Fetching ETF names for Positions sheet...")
            positions_df['etf_name'] = positions_df['ticker'].apply(self._get_etf_name)

            # Reorder columns to put name after ticker
            cols = list(positions_df.columns)
            ticker_idx = cols.index('ticker')
            cols.insert(ticker_idx + 1, cols.pop(cols.index('etf_name')))
            positions_df = positions_df[cols]

            for r_idx, row in enumerate(dataframe_to_rows(positions_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Trades with ETF names
        if len(self.trades_history) > 0:
            ws = wb.create_sheet("Trades")
            trades_df = pd.DataFrame(self.trades_history)
            if 'date' in trades_df.columns:
                trades_df['date'] = pd.to_datetime(trades_df['date']).dt.tz_localize(None)

            # Add ETF names column
            logger.info("Fetching ETF names for Trades sheet...")
            trades_df['etf_name'] = trades_df['ticker'].apply(self._get_etf_name)

            # Reorder columns to put name after ticker
            cols = list(trades_df.columns)
            ticker_idx = cols.index('ticker')
            cols.insert(ticker_idx + 1, cols.pop(cols.index('etf_name')))
            trades_df = trades_df[cols]

            for r_idx, row in enumerate(dataframe_to_rows(trades_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Stop-losses
        if len(self.stop_loss_history) > 0:
            ws = wb.create_sheet("StopLosses")
            sl_df = pd.DataFrame(self.stop_loss_history)
            if 'date' in sl_df.columns:
                sl_df['date'] = pd.to_datetime(sl_df['date']).dt.tz_localize(None)

            for r_idx, row in enumerate(dataframe_to_rows(sl_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(filename)
        logger.info(f"Results saved to {filename}")


def main():
    parser = argparse.ArgumentParser(description="Run stop-loss strategy backtest")
    parser.add_argument("--start", type=str, required=True, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end", type=str, default=datetime.now().strftime("%Y-%m-%d"), help="End date")
    parser.add_argument("--capital", type=float, default=100000, help="Initial capital")
    parser.add_argument("--monthly-add", type=float, default=0, help="Monthly addition")
    parser.add_argument("--positions", type=int, default=20, help="Number of positions")
    # Capital deployment parameters
    parser.add_argument("--immediate-deploy-pct", type=float, default=0.588, help="Immediate deployment pct")
    parser.add_argument("--monthly-deploy", type=float, default=10000, help="Monthly deployment from reserve")
    parser.add_argument("--fixed-income", type=str, default="SGOV", help="Fixed income ticker")
    # Stop-loss parameters
    parser.add_argument("--no-entry-stop", action="store_true", help="Disable entry price stop")
    parser.add_argument("--trailing-stop", action="store_true", help="Enable pure trailing stop from peak")
    parser.add_argument("--trailing-pct", type=float, default=0.10, help="Trailing stop percent from peak e.g. 0.10 for 10 pct")
    parser.add_argument("--trailing-threshold", type=float, default=0.10, help="Gain-triggered trailing stop threshold")
    parser.add_argument("--trailing-distance", type=float, default=0.08, help="Gain-triggered trailing stop distance")
    parser.add_argument("--factor-weeks", type=int, default=2, help="Factor recalc frequency (weeks)")
    parser.add_argument("--output", type=str, default=None, help="Output file")

    args = parser.parse_args()

    if args.output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = RESULTS_DIR / f"stop_loss_backtest_{timestamp}.xlsx"
    else:
        output_file = Path(args.output)

    backtest = StopLossBacktest(
        start_date=args.start,
        end_date=args.end,
        initial_capital=args.capital,
        monthly_addition=args.monthly_add,
        num_positions=args.positions,
        immediate_deploy_pct=args.immediate_deploy_pct,
        monthly_deploy_from_reserve=args.monthly_deploy,
        fixed_income_ticker=args.fixed_income,
        use_entry_price_stop=not args.no_entry_stop,
        use_trailing_stop=args.trailing_stop,
        trailing_stop_pct=args.trailing_pct,
        trailing_stop_threshold=args.trailing_threshold,
        trailing_stop_distance=args.trailing_distance,
        factor_recalc_weeks=args.factor_weeks,
    )

    success = backtest.run()

    if success:
        backtest.save_to_excel(str(output_file))
        results = backtest.generate_results()

        print("\n" + "=" * 80)
        print("STOP-LOSS STRATEGY RESULTS")
        print("=" * 80)
        print(f"Period: {results['start_date'].date()} to {results['end_date'].date()}")
        print(f"Total Return: {results['total_return']:.2%}")
        print(f"CAGR: {results['cagr']:.2%}")
        print(f"Sharpe Ratio: {results['sharpe_ratio']:.2f}")
        print(f"Max Drawdown: {results['max_drawdown']:.2%}")
        print(f"Stop-Losses Triggered: {results['num_stop_losses']}")
        print(f"Avg Stop-Loss: {results['avg_stop_loss']:.2%}")
        print(f"Win Rate: {results['win_rate']:.1%}")
        print(f"Total Trades: {results['total_trades']}")
        print("=" * 80)
        print(f"\nResults saved to: {output_file}")


if __name__ == "__main__":
    main()
