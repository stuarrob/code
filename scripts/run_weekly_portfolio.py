#!/usr/bin/env python3
"""
AQR Multi-Factor ETF Strategy - Weekly Portfolio Automation Script

This script automates the entire weekly portfolio management workflow:
1. Update price data
2. Validate data quality
3. Calculate factor scores
4. Generate portfolio recommendations
5. Check rebalancing need against current positions
6. Update Excel tracking workbook
7. Log all operations

Usage:
    python scripts/run_weekly_portfolio.py --capital 1000000 --positions 20
    python scripts/run_weekly_portfolio.py --capital 1000000 --positions 20 --force-rebalance
"""

import argparse
import logging
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.factors import (
    FactorIntegrator,
    MomentumFactor,
    QualityFactor,
    SimplifiedValueFactor,
    VolatilityFactor,
)
from src.portfolio import (
    MeanVarianceOptimizer,
    MinVarianceOptimizer,
    RankBasedOptimizer,
    StopLossManager,
    ThresholdRebalancer,
)

# Setup logging
LOG_DIR = PROJECT_ROOT / "logs"
LOG_DIR.mkdir(exist_ok=True)
log_file = LOG_DIR / f"weekly_automation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler(log_file), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# Constants
DATA_DIR = PROJECT_ROOT / "data"
RESULTS_DIR = PROJECT_ROOT / "results"
TRACKING_FILE = RESULTS_DIR / "portfolio_tracking.xlsx"


class PortfolioAutomation:
    """Automates weekly portfolio management workflow"""

    def __init__(
        self,
        capital: float,
        num_positions: int,
        optimizer: str = "mvo",
        drift_threshold: float = 0.05,
        force_rebalance: bool = False,
        additional_cash: float = 0.0,
    ):
        self.capital = capital
        self.num_positions = num_positions
        self.optimizer_name = optimizer
        self.drift_threshold = drift_threshold
        self.force_rebalance = force_rebalance
        self.additional_cash = additional_cash

        # Cache for ETF names
        self._etf_names_cache = {}

        # Initialize components
        self.factor_calculators = {
            "momentum": MomentumFactor(lookback=252, skip_recent=21),
            "quality": QualityFactor(lookback=252),
            "value": SimplifiedValueFactor(),
            "volatility": VolatilityFactor(lookback=60),
        }
        self.integrator = FactorIntegrator(
            factor_weights={
                "momentum": 0.25,
                "quality": 0.25,
                "value": 0.25,
                "volatility": 0.25,
            }
        )
        self.rebalancer = ThresholdRebalancer(drift_threshold=drift_threshold)
        self.risk_manager = StopLossManager(position_stop_loss=0.12)

        # Initialize optimizer
        if optimizer == "mvo":
            self.optimizer = MeanVarianceOptimizer(
                num_positions=num_positions,
                lookback=60,
                risk_aversion=1.0,
                axioma_penalty=0.01,
                use_factor_scores_as_alpha=True,
            )
        elif optimizer == "minvar":
            self.optimizer = MinVarianceOptimizer(
                num_positions=num_positions, lookback=60, risk_penalty=0.01
            )
        elif optimizer == "rankbased":
            self.optimizer = RankBasedOptimizer(
                num_positions=num_positions, top_weight=0.08, bottom_weight=0.02
            )
        else:
            raise ValueError(f"Unknown optimizer: {optimizer}")

        logger.info(f"Initialized PortfolioAutomation with {optimizer} optimizer")
        logger.info(
            f"Capital: ${capital:,.0f}, Positions: {num_positions}, Drift: {drift_threshold:.1%}"
        )
        if additional_cash > 0:
            logger.info(f"Additional cash to invest: ${additional_cash:,.2f}")

    def _get_etf_names(self, tickers: list) -> dict:
        """Fetch ETF long names from yfinance"""
        names = {}
        for ticker in tickers:
            if ticker in self._etf_names_cache:
                names[ticker] = self._etf_names_cache[ticker]
                continue

            try:
                etf = yf.Ticker(ticker)
                info = etf.info
                long_name = info.get('longName') or info.get('shortName') or ticker
                names[ticker] = long_name
                self._etf_names_cache[ticker] = long_name
            except:
                # If fetch fails, use ticker as name
                names[ticker] = ticker
                self._etf_names_cache[ticker] = ticker

        return names

    def step1_update_data(self) -> bool:
        """Step 1: Update price data"""
        logger.info("=" * 80)
        logger.info("STEP 1: Updating price data")
        logger.info("=" * 80)

        try:
            result = subprocess.run(
                [
                    "python",
                    str(PROJECT_ROOT / "scripts" / "01_collect_universe.py"),
                ],
                capture_output=True,
                text=True,
                timeout=900,  # 15 minutes
            )

            if result.returncode != 0:
                logger.error(f"Data collection failed: {result.stderr}")
                return False

            logger.info("Price data updated successfully")
            return True

        except subprocess.TimeoutExpired:
            logger.error("Data collection timed out after 15 minutes")
            return False
        except Exception as e:
            logger.error(f"Error updating data: {e}")
            return False

    def step2_validate_data(self) -> bool:
        """Step 2: Validate data quality"""
        logger.info("=" * 80)
        logger.info("STEP 2: Validating data quality")
        logger.info("=" * 80)

        try:
            result = subprocess.run(
                ["python", str(PROJECT_ROOT / "scripts" / "validate_real_data.py")],
                capture_output=True,
                text=True,
                timeout=180,  # 3 minutes
            )

            if result.returncode != 0:
                logger.error(f"Data validation failed: {result.stderr}")
                return False

            # Check if filtered data exists
            filtered_file = DATA_DIR / "processed" / "etf_prices_filtered.parquet"
            if not filtered_file.exists():
                logger.error("Filtered data file not created")
                return False

            # Check number of ETFs
            prices = pd.read_parquet(filtered_file)
            num_etfs = len(prices.columns)
            logger.info(f"Data validated: {num_etfs} eligible ETFs")

            if num_etfs < 500:
                logger.warning(f"Only {num_etfs} ETFs available (expected ~620)")

            return True

        except Exception as e:
            logger.error(f"Error validating data: {e}")
            return False

    def step3_calculate_factors(self) -> pd.Series:
        """Step 3: Calculate factor scores"""
        logger.info("=" * 80)
        logger.info("STEP 3: Calculating factor scores")
        logger.info("=" * 80)

        try:
            # Load filtered prices
            prices = pd.read_parquet(
                DATA_DIR / "processed" / "etf_prices_filtered.parquet"
            )
            logger.info(f"Loaded prices for {len(prices.columns)} ETFs")

            # Calculate individual factor scores
            factor_scores_dict = {}
            for factor_name, calculator in self.factor_calculators.items():
                logger.info(f"Calculating {factor_name} scores...")

                # Value factor needs expense ratios
                if factor_name == "value":
                    # Generate synthetic expense ratios (replace with real data in production)
                    expense_ratios = pd.Series(
                        np.random.uniform(0.0005, 0.01, len(prices.columns)),
                        index=prices.columns,
                    )
                    scores = calculator.calculate(prices, expense_ratios)
                else:
                    scores = calculator.calculate(prices)

                valid_count = scores.notna().sum()
                logger.info(f"  {factor_name}: {valid_count} valid scores")
                factor_scores_dict[factor_name] = scores

            # Combine into DataFrame for integrator
            factor_scores_df = pd.DataFrame(factor_scores_dict)

            # Integrate factors
            logger.info("Integrating factor scores...")
            integrated_scores = self.integrator.integrate(factor_scores_df)

            valid_integrated = integrated_scores.notna().sum()
            logger.info(
                f"Final integrated scores: {valid_integrated} ETFs with valid scores"
            )

            # Save factor scores
            scores_file = DATA_DIR / "factor_scores_latest.parquet"
            integrated_scores.to_frame("integrated_score").to_parquet(scores_file)
            logger.info(f"Saved factor scores to {scores_file}")

            return integrated_scores

        except Exception as e:
            logger.error(f"Error calculating factors: {e}")
            raise

    def step4_generate_portfolio(self, factor_scores: pd.Series) -> tuple:
        """Step 4: Generate target portfolio"""
        logger.info("=" * 80)
        logger.info("STEP 4: Generating target portfolio")
        logger.info("=" * 80)

        try:
            # Load prices for covariance calculation
            prices = pd.read_parquet(
                DATA_DIR / "processed" / "etf_prices_filtered.parquet"
            )
            returns = prices.pct_change(fill_method=None).dropna()

            # Filter to ETFs with valid scores
            valid_scores = factor_scores.dropna()
            valid_tickers = valid_scores.index.intersection(returns.columns)

            logger.info(f"Optimizing portfolio from {len(valid_tickers)} ETFs")

            # Optimize (different APIs for different optimizers)
            if self.optimizer_name in ["mvo", "minvar"]:
                weights = self.optimizer.optimize(
                    valid_scores[valid_tickers], prices[valid_tickers]
                )
            else:  # rankbased
                weights = self.optimizer.optimize(valid_scores[valid_tickers])

            # Convert to dollar positions
            positions = (weights * self.capital).round(2)

            # Get current prices
            latest_prices = prices.iloc[-1][positions.index]
            shares = (positions / latest_prices).round(0)

            # Fetch ETF long names
            logger.info("Fetching ETF names...")
            etf_names = self._get_etf_names(list(positions.index))

            # Create target portfolio DataFrame
            target_portfolio = pd.DataFrame(
                {
                    "ticker": positions.index,
                    "name": [etf_names[ticker] for ticker in positions.index],
                    "weight": weights[positions.index],
                    "value": positions[positions.index],
                    "shares": shares[positions.index],
                    "price": latest_prices[positions.index],
                    "factor_score": valid_scores[positions.index],
                }
            )

            logger.info(f"Generated portfolio with {len(target_portfolio)} positions")
            logger.info(f"Total value: ${target_portfolio['value'].sum():,.2f}")

            # Calculate expected portfolio metrics
            # Get the tickers that have weights (the selected positions)
            selected_tickers = weights.index

            # Calculate expected return using selected positions
            port_return = (weights * valid_scores[selected_tickers]).sum()

            # Calculate portfolio volatility using only selected positions
            selected_cov = returns[selected_tickers].cov()
            port_vol = np.sqrt(
                (weights.values @ selected_cov.values @ weights.values) * 252
            )
            port_sharpe = port_return / port_vol if port_vol > 0 else 0

            logger.info(f"Expected return: {port_return:.2%}")
            logger.info(f"Expected volatility: {port_vol:.2%}")
            logger.info(f"Expected Sharpe: {port_sharpe:.2f}")

            metrics = {
                "expected_return": port_return,
                "expected_volatility": port_vol,
                "expected_sharpe": port_sharpe,
            }

            return target_portfolio, metrics

        except Exception as e:
            logger.error(f"Error generating portfolio: {e}")
            raise

    def step5_check_rebalancing(self, target_portfolio: pd.DataFrame) -> tuple:
        """Step 5: Check if rebalancing is needed"""
        logger.info("=" * 80)
        logger.info("STEP 5: Checking rebalancing need")
        logger.info("=" * 80)

        try:
            # Load current portfolio from Excel
            current_portfolio = self._load_current_portfolio()

            if current_portfolio is None or len(current_portfolio) == 0:
                logger.info("No current portfolio found - initial portfolio setup")
                trades = self._generate_initial_trades(target_portfolio)
                return True, trades, None

            logger.info(f"Current portfolio: {len(current_portfolio)} positions")

            # Check if rebalancing needed
            needs_rebalance = self.rebalancer.needs_rebalancing(
                current_weights=current_portfolio.set_index("ticker")["weight"],
                target_weights=target_portfolio.set_index("ticker")["weight"],
            )

            if self.force_rebalance:
                logger.info(
                    "Force rebalance flag set - rebalancing regardless of drift"
                )
                needs_rebalance = True

            if needs_rebalance:
                logger.info(f"Rebalancing needed (drift > {self.drift_threshold:.1%})")
                trades = self._generate_rebalancing_trades(
                    current_portfolio, target_portfolio
                )
                return True, trades, current_portfolio
            else:
                logger.info(
                    f"No rebalancing needed (drift < {self.drift_threshold:.1%})"
                )
                return False, None, current_portfolio

        except Exception as e:
            logger.error(f"Error checking rebalancing: {e}")
            raise

    def step6_update_tracking(
        self,
        target_portfolio: pd.DataFrame,
        trades: pd.DataFrame,
        needs_rebalance: bool,
        metrics: dict,
    ):
        """Step 6: Update Excel tracking workbook"""
        logger.info("=" * 80)
        logger.info("STEP 6: Updating Excel tracking workbook")
        logger.info("=" * 80)

        try:
            # Initialize or load workbook
            if not TRACKING_FILE.exists():
                logger.info("Creating new tracking workbook")
                wb = self._create_tracking_workbook()
            else:
                logger.info("Loading existing tracking workbook")
                wb = load_workbook(TRACKING_FILE)

            timestamp = datetime.now()

            # Update Positions sheet
            self._update_positions_sheet(wb, target_portfolio, timestamp)

            # Update Trades sheet (if rebalancing)
            if needs_rebalance and trades is not None:
                self._update_trades_sheet(wb, trades, timestamp)

            # Update Performance sheet
            self._update_performance_sheet(wb, target_portfolio, metrics, timestamp)

            # Update Metadata sheet
            self._update_metadata_sheet(wb, needs_rebalance, timestamp)

            # Save workbook
            wb.save(TRACKING_FILE)
            logger.info(f"Tracking workbook updated: {TRACKING_FILE}")

        except Exception as e:
            logger.error(f"Error updating tracking: {e}")
            raise

    def _load_current_portfolio(self) -> pd.DataFrame:
        """Load current portfolio from Excel"""
        if not TRACKING_FILE.exists():
            return None

        try:
            df = pd.read_excel(TRACKING_FILE, sheet_name="Positions")
            if len(df) == 0:
                return None

            # Get most recent positions
            if "timestamp" in df.columns:
                latest = df["timestamp"].max()
                current = df[df["timestamp"] == latest].copy()
            else:
                current = df.copy()

            logger.info(f"Loaded current portfolio: {len(current)} positions")
            return current

        except Exception as e:
            logger.warning(f"Could not load current portfolio: {e}")
            return None

    def _generate_initial_trades(self, target_portfolio: pd.DataFrame) -> pd.DataFrame:
        """Generate trades for initial portfolio setup"""
        trades = pd.DataFrame(
            {
                "ticker": target_portfolio["ticker"],
                "name": target_portfolio["name"],
                "action": "BUY",
                "shares": target_portfolio["shares"],
                "price": target_portfolio["price"],
                "value": target_portfolio["value"],
            }
        )

        logger.info(f"Generated {len(trades)} initial buy trades")
        return trades

    def _generate_rebalancing_trades(
        self, current: pd.DataFrame, target: pd.DataFrame
    ) -> pd.DataFrame:
        """Generate trades to rebalance from current to target portfolio"""
        trades_list = []

        # Get current prices
        prices = pd.read_parquet(DATA_DIR / "processed" / "etf_prices_filtered.parquet")
        latest_prices = prices.iloc[-1]

        current_dict = current.set_index("ticker")["shares"].to_dict()
        target_dict = target.set_index("ticker")["shares"].to_dict()

        # Build name lookup from both current and target
        name_dict = {}
        if 'name' in current.columns:
            name_dict.update(current.set_index("ticker")["name"].to_dict())
        if 'name' in target.columns:
            name_dict.update(target.set_index("ticker")["name"].to_dict())

        all_tickers = set(current_dict.keys()) | set(target_dict.keys())

        # Get names for any tickers we don't have
        missing_tickers = [t for t in all_tickers if t not in name_dict]
        if missing_tickers:
            new_names = self._get_etf_names(missing_tickers)
            name_dict.update(new_names)

        for ticker in all_tickers:
            current_shares = current_dict.get(ticker, 0)
            target_shares = target_dict.get(ticker, 0)
            diff = target_shares - current_shares

            if abs(diff) < 1:  # Skip tiny differences
                continue

            price = latest_prices.get(ticker, 0)
            if price == 0:
                logger.warning(f"No price for {ticker}, skipping")
                continue

            action = "BUY" if diff > 0 else "SELL"
            trades_list.append(
                {
                    "ticker": ticker,
                    "name": name_dict.get(ticker, ticker),
                    "action": action,
                    "shares": abs(diff),
                    "price": price,
                    "value": abs(diff) * price,
                }
            )

        trades = pd.DataFrame(trades_list)
        logger.info(f"Generated {len(trades)} rebalancing trades")
        logger.info(f"  BUY: {(trades['action'] == 'BUY').sum()} trades")
        logger.info(f"  SELL: {(trades['action'] == 'SELL').sum()} trades")

        return trades

    def _create_tracking_workbook(self) -> Workbook:
        """Create new Excel tracking workbook with all sheets"""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        wb.create_sheet("Positions")
        wb.create_sheet("Trades")
        wb.create_sheet("Performance")
        wb.create_sheet("Metadata")

        # Format headers
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.sheet_view.showGridLines = True

        return wb

    def _update_positions_sheet(
        self, wb: Workbook, portfolio: pd.DataFrame, timestamp: datetime
    ):
        """Update Positions sheet in workbook"""
        ws = wb["Positions"]

        # Add timestamp column
        portfolio_with_ts = portfolio.copy()
        portfolio_with_ts.insert(0, "timestamp", timestamp)

        # Append to sheet
        if ws.max_row == 1:
            # First entry - add headers
            for r_idx, row in enumerate(
                dataframe_to_rows(portfolio_with_ts, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        else:
            # Append data
            start_row = ws.max_row + 1
            for r_idx, row in enumerate(
                dataframe_to_rows(portfolio_with_ts, index=False, header=False),
                start_row,
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        logger.info(f"Updated Positions sheet: {len(portfolio)} positions")

    def _update_trades_sheet(
        self, wb: Workbook, trades: pd.DataFrame, timestamp: datetime
    ):
        """Update Trades sheet in workbook"""
        ws = wb["Trades"]

        # Add timestamp and execution status
        trades_with_ts = trades.copy()
        trades_with_ts.insert(0, "timestamp", timestamp)
        trades_with_ts["executed"] = False  # Mark as not yet executed

        # Append to sheet
        if ws.max_row == 1:
            # First entry - add headers
            for r_idx, row in enumerate(
                dataframe_to_rows(trades_with_ts, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        else:
            # Append data
            start_row = ws.max_row + 1
            for r_idx, row in enumerate(
                dataframe_to_rows(trades_with_ts, index=False, header=False), start_row
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        logger.info(f"Updated Trades sheet: {len(trades)} trades")

    def _update_performance_sheet(
        self, wb: Workbook, portfolio: pd.DataFrame, metrics: dict, timestamp: datetime
    ):
        """Update Performance sheet in workbook"""
        ws = wb["Performance"]

        # Calculate current metrics
        total_value = portfolio["value"].sum()
        num_positions = len(portfolio)

        perf_data = {
            "timestamp": timestamp,
            "total_value": total_value,
            "num_positions": num_positions,
            "expected_return": metrics.get("expected_return", 0),
            "expected_volatility": metrics.get("expected_volatility", 0),
            "expected_sharpe": metrics.get("expected_sharpe", 0),
        }

        # Append to sheet
        if ws.max_row == 1:
            # First entry - add headers
            headers = list(perf_data.keys())
            for c_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=c_idx, value=header)
            for c_idx, value in enumerate(perf_data.values(), 1):
                ws.cell(row=2, column=c_idx, value=value)
        else:
            # Append data
            row_idx = ws.max_row + 1
            for c_idx, value in enumerate(perf_data.values(), 1):
                ws.cell(row=row_idx, column=c_idx, value=value)

        logger.info(f"Updated Performance sheet")

    def _update_metadata_sheet(
        self, wb: Workbook, needs_rebalance: bool, timestamp: datetime
    ):
        """Update Metadata sheet in workbook"""
        ws = wb["Metadata"]

        metadata = {
            "timestamp": timestamp,
            "rebalanced": needs_rebalance,
            "optimizer": self.optimizer_name,
            "drift_threshold": self.drift_threshold,
            "capital": self.capital,
            "num_positions": self.num_positions,
        }

        # Append to sheet
        if ws.max_row == 1:
            # First entry - add headers
            headers = list(metadata.keys())
            for c_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=c_idx, value=header)
            for c_idx, value in enumerate(metadata.values(), 1):
                ws.cell(row=2, column=c_idx, value=value)
        else:
            # Append data
            row_idx = ws.max_row + 1
            for c_idx, value in enumerate(metadata.values(), 1):
                ws.cell(row=row_idx, column=c_idx, value=value)

        logger.info(f"Updated Metadata sheet")

    def run(self):
        """Run complete weekly automation workflow"""
        logger.info("=" * 80)
        logger.info("STARTING WEEKLY PORTFOLIO AUTOMATION")
        logger.info(f"Time: {datetime.now()}")
        logger.info("=" * 80)

        try:
            # Step 1: Update data
            if not self.step1_update_data():
                logger.error("Data update failed - aborting")
                return False

            # Step 2: Validate data
            if not self.step2_validate_data():
                logger.error("Data validation failed - aborting")
                return False

            # Check if we need to adjust capital for additional cash
            if self.additional_cash > 0:
                current_portfolio = self._load_current_portfolio()
                if current_portfolio is not None and len(current_portfolio) > 0:
                    current_value = current_portfolio['value'].sum()
                    new_capital = current_value + self.additional_cash
                    logger.info(f"Adjusting capital from ${self.capital:,.2f} to ${new_capital:,.2f} (added ${self.additional_cash:,.2f})")
                    self.capital = new_capital
                    self.force_rebalance = True  # Force rebalance when adding cash

            # Step 3: Calculate factors
            factor_scores = self.step3_calculate_factors()

            # Step 4: Generate portfolio
            target_portfolio, metrics = self.step4_generate_portfolio(factor_scores)

            # Step 5: Check rebalancing
            needs_rebalance, trades, current_portfolio = self.step5_check_rebalancing(
                target_portfolio
            )

            # Step 6: Update tracking
            self.step6_update_tracking(
                target_portfolio, trades, needs_rebalance, metrics
            )

            # Summary
            logger.info("=" * 80)
            logger.info("WEEKLY AUTOMATION COMPLETED SUCCESSFULLY")
            logger.info("=" * 80)
            logger.info(f"Rebalancing needed: {needs_rebalance}")
            if needs_rebalance and trades is not None:
                logger.info(f"Trades to execute: {len(trades)}")
                logger.info(f"  BUY orders: {(trades['action'] == 'BUY').sum()}")
                logger.info(f"  SELL orders: {(trades['action'] == 'SELL').sum()}")
            logger.info(f"Tracking workbook: {TRACKING_FILE}")
            logger.info(f"Log file: {log_file}")

            return True

        except Exception as e:
            logger.error(f"Automation failed: {e}", exc_info=True)
            return False


def main():
    parser = argparse.ArgumentParser(
        description="Automate weekly portfolio management workflow"
    )
    parser.add_argument(
        "--capital",
        type=float,
        default=1000000,
        help="Portfolio capital (default: 1000000)",
    )
    parser.add_argument(
        "--positions", type=int, default=20, help="Number of positions (default: 20)"
    )
    parser.add_argument(
        "--optimizer",
        choices=["mvo", "minvar", "rankbased"],
        default="mvo",
        help="Optimizer to use (default: mvo)",
    )
    parser.add_argument(
        "--drift-threshold",
        type=float,
        default=0.05,
        help="Rebalancing drift threshold (default: 0.05)",
    )
    parser.add_argument(
        "--force-rebalance",
        action="store_true",
        help="Force rebalancing regardless of drift",
    )
    parser.add_argument(
        "--additional-cash",
        type=float,
        default=0.0,
        help="Additional cash to invest (e.g., monthly savings)",
    )

    args = parser.parse_args()

    # Create automation instance
    automation = PortfolioAutomation(
        capital=args.capital,
        num_positions=args.positions,
        optimizer=args.optimizer,
        drift_threshold=args.drift_threshold,
        force_rebalance=args.force_rebalance,
        additional_cash=args.additional_cash,
    )

    # Run automation
    success = automation.run()

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
