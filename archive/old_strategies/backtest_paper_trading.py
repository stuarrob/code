#!/usr/bin/env python3
"""
Historical Paper Trading Backtest

Simulates weekly portfolio management from a start date to end date,
with optional monthly capital additions. Generates a spreadsheet with
performance metrics.

Usage:
    python scripts/backtest_paper_trading.py --start 2025-03-01 --end 2025-10-31 --capital 100000 --monthly-add 5000
"""

import argparse
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.factors import (
    FactorIntegrator,
    MomentumFactor,
    QualityFactor,
    SimplifiedValueFactor,
    VolatilityFactor,
)
from src.portfolio import (
    MeanVarianceOptimizer,
)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# Constants
DATA_DIR = PROJECT_ROOT / "data"
RESULTS_DIR = PROJECT_ROOT / "results"


class PaperTradingBacktest:
    """Simulates weekly paper trading with performance tracking"""

    def __init__(
        self,
        start_date: str,
        end_date: str,
        initial_capital: float,
        monthly_addition: float,
        num_positions: int = 20,
        rebalance_freq: str = "weekly",
        factor_ema_alpha: float = 0.3,  # NEW: Factor smoothing
        drift_threshold: float = 0.15,  # NEW: Rebalancing threshold
        use_real_expense_ratios: bool = True,  # NEW: Fix value factor
    ):
        self.start_date = pd.Timestamp(start_date)
        self.end_date = pd.Timestamp(end_date)
        self.initial_capital = initial_capital
        self.monthly_addition = monthly_addition
        self.num_positions = num_positions
        self.rebalance_freq = rebalance_freq

        # Stability parameters
        self.factor_ema_alpha = factor_ema_alpha
        self.drift_threshold = drift_threshold
        self.use_real_expense_ratios = use_real_expense_ratios

        # Results tracking
        self.portfolio_history = []
        self.trades_history = []
        self.performance_history = []

        # State for factor smoothing
        self.smoothed_factor_scores = None  # NEW: Track smoothed scores

        # Cache for ETF names
        self._etf_names_cache = {}

        # Cache for expense ratios
        self._expense_ratio_cache = {}  # NEW

        # Initialize components
        self.factor_calculators = {
            "momentum": MomentumFactor(lookback=252, skip_recent=21),
            "quality": QualityFactor(lookback=252),
            "value": SimplifiedValueFactor(),
            "volatility": VolatilityFactor(lookback=60),
        }
        self.integrator = FactorIntegrator(
            factor_weights={
                "momentum": 0.25,
                "quality": 0.25,
                "value": 0.25,
                "volatility": 0.25,
            }
        )
        self.optimizer = MeanVarianceOptimizer(
            num_positions=num_positions,
            lookback=60,
            risk_aversion=1.0,
            axioma_penalty=0.01,
            use_factor_scores_as_alpha=True,
        )

        logger.info(f"Initialized backtest from {start_date} to {end_date}")
        logger.info(f"Initial capital: ${initial_capital:,.0f}")
        logger.info(f"Monthly addition: ${monthly_addition:,.0f}")
        logger.info(f"Stability settings:")
        logger.info(f"  - Factor EMA alpha: {factor_ema_alpha} (lower = more smoothing)")
        logger.info(f"  - Drift threshold: {drift_threshold:.1%}")
        logger.info(f"  - Real expense ratios: {use_real_expense_ratios}")

    def _get_etf_names(self, tickers: list) -> dict:
        """Fetch ETF long names from yfinance"""
        names = {}
        for ticker in tickers:
            if ticker in self._etf_names_cache:
                names[ticker] = self._etf_names_cache[ticker]
                continue

            try:
                etf = yf.Ticker(ticker)
                info = etf.info
                long_name = info.get("longName") or info.get("shortName") or ticker
                names[ticker] = long_name
                self._etf_names_cache[ticker] = long_name
            except:
                names[ticker] = ticker
                self._etf_names_cache[ticker] = ticker

        return names

    def _get_expense_ratios(self, tickers: list) -> pd.Series:
        """Fetch real expense ratios from yfinance with caching"""
        expense_ratios = {}

        for ticker in tickers:
            if ticker in self._expense_ratio_cache:
                expense_ratios[ticker] = self._expense_ratio_cache[ticker]
                continue

            try:
                etf = yf.Ticker(ticker)
                info = etf.info
                # Expense ratio is typically stored as a decimal (e.g., 0.0003 for 0.03%)
                er = info.get('expenseRatio', None)
                if er is None or pd.isna(er):
                    # Default to median expense ratio if not available
                    er = 0.005  # 0.5%
                expense_ratios[ticker] = float(er)
                self._expense_ratio_cache[ticker] = float(er)
            except Exception as e:
                # Default to median expense ratio on error
                expense_ratios[ticker] = 0.005
                self._expense_ratio_cache[ticker] = 0.005

        return pd.Series(expense_ratios)

    def _get_rebalance_dates(self, prices: pd.DataFrame) -> list:
        """Generate list of rebalancing dates"""
        # Make start/end dates timezone-aware if prices index has timezone
        start_date = self.start_date
        end_date = self.end_date

        if prices.index.tz is not None:
            if start_date.tz is None:
                start_date = start_date.tz_localize(prices.index.tz)
            if end_date.tz is None:
                end_date = end_date.tz_localize(prices.index.tz)

        # Get all trading days between start and end
        trading_days = prices.index[
            (prices.index >= start_date) & (prices.index <= end_date)
        ]

        if len(trading_days) == 0:
            raise ValueError(f"No trading days found between {self.start_date} and {self.end_date}")

        if self.rebalance_freq == "weekly":
            # Get first trading day of each week (Monday or first day if Monday is holiday)
            rebalance_dates = []
            current_week = None

            for date in trading_days:
                week = date.isocalendar()[1]
                year = date.isocalendar()[0]
                week_key = (year, week)

                if week_key != current_week:
                    rebalance_dates.append(date)
                    current_week = week_key

            return rebalance_dates
        else:
            raise ValueError(f"Unsupported rebalance frequency: {self.rebalance_freq}")

    def _is_last_week_of_month(self, date: pd.Timestamp) -> bool:
        """Check if date is in the last week of the month"""
        # Get the last day of the month
        if date.month == 12:
            next_month = date.replace(year=date.year + 1, month=1, day=1)
        else:
            next_month = date.replace(month=date.month + 1, day=1)

        last_day = next_month - timedelta(days=1)

        # Check if date is within 7 days of month end
        days_to_end = (last_day - date).days
        return days_to_end < 7

    def _calculate_factors_at_date(
        self, prices: pd.DataFrame, date: pd.Timestamp
    ) -> pd.Series:
        """Calculate factor scores using data up to given date with smoothing"""
        # Get prices up to this date
        historical_prices = prices.loc[:date]

        # Need at least 252 days for momentum calculation
        if len(historical_prices) < 252:
            logger.warning(f"Insufficient history at {date.date()} ({len(historical_prices)} days)")
            return pd.Series(dtype=float)

        # Calculate individual factor scores
        factor_scores_dict = {}
        for factor_name, calculator in self.factor_calculators.items():
            try:
                if factor_name == "value":
                    if self.use_real_expense_ratios:
                        # Use real expense ratios
                        tickers = list(historical_prices.columns)
                        expense_ratios = self._get_expense_ratios(tickers)
                    else:
                        # Generate synthetic expense ratios (old behavior)
                        expense_ratios = pd.Series(
                            np.random.uniform(0.0005, 0.01, len(historical_prices.columns)),
                            index=historical_prices.columns,
                        )
                    scores = calculator.calculate(historical_prices, expense_ratios)
                else:
                    scores = calculator.calculate(historical_prices)

                factor_scores_dict[factor_name] = scores
            except Exception as e:
                logger.warning(f"Error calculating {factor_name} at {date.date()}: {e}")
                factor_scores_dict[factor_name] = pd.Series(dtype=float)

        # Combine into DataFrame
        raw_factor_scores = pd.DataFrame(factor_scores_dict)

        # Apply exponential moving average smoothing
        if self.smoothed_factor_scores is not None:
            # EMA: smoothed = alpha * current + (1 - alpha) * previous
            alpha = self.factor_ema_alpha
            smoothed = alpha * raw_factor_scores + (1 - alpha) * self.smoothed_factor_scores

            # Handle new ETFs (no previous score)
            smoothed = smoothed.fillna(raw_factor_scores)

            # Handle removed ETFs (update with current universe)
            smoothed = smoothed.reindex(raw_factor_scores.index, fill_value=np.nan)
            smoothed = smoothed.fillna(raw_factor_scores)
        else:
            # First iteration - no smoothing
            smoothed = raw_factor_scores

        # Store smoothed scores for next iteration
        self.smoothed_factor_scores = smoothed.copy()

        # Integrate factors
        integrated_scores = self.integrator.integrate(smoothed)

        return integrated_scores

    def _optimize_portfolio(
        self, factor_scores: pd.Series, prices: pd.DataFrame, date: pd.Timestamp, capital: float
    ) -> pd.DataFrame:
        """Generate optimal portfolio for given date and capital"""
        # Get historical prices up to this date
        historical_prices = prices.loc[:date]

        # Filter to ETFs with valid scores
        valid_scores = factor_scores.dropna()
        valid_tickers = valid_scores.index.intersection(historical_prices.columns)

        if len(valid_tickers) == 0:
            logger.warning(f"No valid tickers at {date.date()}")
            return pd.DataFrame()

        # Optimize
        try:
            weights = self.optimizer.optimize(
                valid_scores[valid_tickers], historical_prices[valid_tickers]
            )
        except Exception as e:
            logger.error(f"Optimization failed at {date.date()}: {e}")
            return pd.DataFrame()

        # Convert to dollar positions
        positions = (weights * capital).round(2)

        # Get current prices
        latest_prices = historical_prices.iloc[-1][positions.index]
        shares = (positions / latest_prices).round(0)

        # Fetch ETF names
        etf_names = self._get_etf_names(list(positions.index))

        # Create portfolio DataFrame
        portfolio = pd.DataFrame(
            {
                "ticker": positions.index,
                "name": [etf_names[ticker] for ticker in positions.index],
                "weight": weights[positions.index],
                "value": positions[positions.index],
                "shares": shares[positions.index],
                "price": latest_prices[positions.index],
                "factor_score": valid_scores[positions.index],
            }
        )

        return portfolio

    def _calculate_portfolio_value(
        self, portfolio: pd.DataFrame, prices: pd.DataFrame, date: pd.Timestamp
    ) -> float:
        """Calculate portfolio value at given date using current prices"""
        if len(portfolio) == 0:
            return 0.0

        # Find the date in the prices index (handle timezone)
        if date not in prices.index:
            # Try to find closest date
            available_dates = prices.index[prices.index <= date]
            if len(available_dates) == 0:
                logger.warning(f"No prices available for {date.date()}")
                return 0.0
            date = available_dates[-1]

        current_prices = prices.loc[date]
        total_value = 0.0

        for _, row in portfolio.iterrows():
            ticker = row["ticker"]
            shares = row["shares"]
            if ticker in current_prices.index and not pd.isna(current_prices[ticker]):
                total_value += shares * current_prices[ticker]

        return total_value

    def _generate_trades(
        self, old_portfolio: pd.DataFrame, new_portfolio: pd.DataFrame, date: pd.Timestamp
    ) -> pd.DataFrame:
        """Generate trades to rebalance portfolio"""
        trades_list = []

        # Handle initial portfolio (all buys)
        if len(old_portfolio) == 0:
            for _, row in new_portfolio.iterrows():
                if row["shares"] > 0:
                    trades_list.append(
                        {
                            "date": date,
                            "ticker": row["ticker"],
                            "name": row["name"],
                            "action": "BUY",
                            "shares": row["shares"],
                            "price": row["price"],
                            "value": row["value"],
                        }
                    )
        else:
            # Rebalancing trades
            old_dict = old_portfolio.set_index("ticker")["shares"].to_dict()
            new_dict = new_portfolio.set_index("ticker")["shares"].to_dict()
            name_dict = new_portfolio.set_index("ticker")["name"].to_dict()
            price_dict = new_portfolio.set_index("ticker")["price"].to_dict()

            all_tickers = set(old_dict.keys()) | set(new_dict.keys())

            for ticker in all_tickers:
                old_shares = old_dict.get(ticker, 0)
                new_shares = new_dict.get(ticker, 0)
                diff = new_shares - old_shares

                if abs(diff) < 1:
                    continue

                price = price_dict.get(ticker, 0)
                if price == 0:
                    continue

                action = "BUY" if diff > 0 else "SELL"
                trades_list.append(
                    {
                        "date": date,
                        "ticker": ticker,
                        "name": name_dict.get(ticker, ticker),
                        "action": action,
                        "shares": abs(diff),
                        "price": price,
                        "value": abs(diff) * price,
                    }
                )

        return pd.DataFrame(trades_list)

    def run(self):
        """Run the complete backtest"""
        logger.info("=" * 80)
        logger.info("STARTING PAPER TRADING BACKTEST")
        logger.info("=" * 80)

        # Load price data
        logger.info("Loading price data...")
        prices_file = DATA_DIR / "processed" / "etf_prices_filtered.parquet"
        if not prices_file.exists():
            logger.error(f"Price data not found: {prices_file}")
            logger.error("Please run scripts/01_collect_universe.py first")
            return False

        prices = pd.read_parquet(prices_file)
        logger.info(f"Loaded {len(prices.columns)} ETFs, {len(prices)} days of data")
        logger.info(f"Price data range: {prices.index[0].date()} to {prices.index[-1].date()}")

        # Generate rebalancing dates
        rebalance_dates = self._get_rebalance_dates(prices)
        logger.info(f"Generated {len(rebalance_dates)} rebalancing dates")
        logger.info(f"First rebalance: {rebalance_dates[0].date()}")
        logger.info(f"Last rebalance: {rebalance_dates[-1].date()}")

        # Initialize tracking
        current_portfolio = pd.DataFrame()
        current_capital = self.initial_capital
        total_contributions = self.initial_capital

        # Run backtest
        for i, date in enumerate(rebalance_dates):
            logger.info("=" * 80)
            logger.info(f"Rebalance {i+1}/{len(rebalance_dates)}: {date.date()}")
            logger.info("=" * 80)

            # Track if capital was added this period
            capital_added_this_period = 0.0

            # If we have existing portfolio, calculate current value
            if len(current_portfolio) > 0:
                current_value = self._calculate_portfolio_value(
                    current_portfolio, prices, date
                )
                logger.info(f"Current portfolio value: ${current_value:,.2f}")
                current_capital = current_value

            # Check if we should add capital (last week of month)
            if i > 0 and self._is_last_week_of_month(date) and self.monthly_addition > 0:
                current_capital += self.monthly_addition
                total_contributions += self.monthly_addition
                capital_added_this_period = self.monthly_addition
                logger.info(f"Adding ${self.monthly_addition:,.2f} monthly contribution")
                logger.info(f"New capital: ${current_capital:,.2f}")

            # Calculate factor scores
            logger.info("Calculating factor scores...")
            factor_scores = self._calculate_factors_at_date(prices, date)
            valid_count = factor_scores.notna().sum()
            logger.info(f"Valid factor scores: {valid_count}")

            if valid_count == 0:
                logger.warning("No valid factor scores, skipping this date")
                continue

            # Optimize portfolio
            logger.info(f"Optimizing portfolio with ${current_capital:,.2f}...")
            new_portfolio = self._optimize_portfolio(
                factor_scores, prices, date, current_capital
            )

            if len(new_portfolio) == 0:
                logger.warning("Optimization failed, skipping this date")
                continue

            logger.info(f"Generated target portfolio: {len(new_portfolio)} positions")

            # Check if rebalancing is needed (drift threshold)
            needs_rebalance = True  # Default for first portfolio
            if len(current_portfolio) > 0:
                # Calculate drift
                current_weights = current_portfolio.set_index("ticker")["weight"]
                target_weights = new_portfolio.set_index("ticker")["weight"]

                # Align weights (fill missing with 0)
                all_tickers = set(current_weights.index) | set(target_weights.index)
                current_aligned = pd.Series({t: current_weights.get(t, 0) for t in all_tickers})
                target_aligned = pd.Series({t: target_weights.get(t, 0) for t in all_tickers})

                # Calculate total drift
                total_drift = (current_aligned - target_aligned).abs().sum()
                needs_rebalance = total_drift > self.drift_threshold

                logger.info(f"Portfolio drift: {total_drift:.2%} (threshold: {self.drift_threshold:.1%})")

                if not needs_rebalance:
                    logger.info(f"✓ SKIPPING REBALANCE - Drift below threshold")
                    # Keep current portfolio, but update value tracking
                    # Store current portfolio with updated timestamp for tracking
                    portfolio_with_date = current_portfolio.copy()
                    portfolio_with_date["date"] = date
                    self.portfolio_history.append(portfolio_with_date)

                    # Track performance without trades
                    portfolio_value = self._calculate_portfolio_value(current_portfolio, prices, date)

                    if i == 0:
                        period_return = 0.0
                        cumulative_return = 0.0
                    else:
                        prev_data = self.performance_history[-1]
                        prev_portfolio_value = prev_data["portfolio_value"]
                        period_return = (portfolio_value - prev_portfolio_value - capital_added_this_period) / prev_portfolio_value if prev_portfolio_value > 0 else 0
                        cumulative_return = (portfolio_value - total_contributions) / total_contributions

                    self.performance_history.append({
                        "date": date,
                        "portfolio_value": portfolio_value,
                        "contributions": total_contributions,
                        "period_return": period_return,
                        "cumulative_return": cumulative_return,
                        "num_positions": len(current_portfolio),
                    })

                    logger.info(f"Portfolio value: ${portfolio_value:,.2f}")
                    logger.info(f"Period return: {period_return:+.2%}")
                    logger.info(f"Cumulative return: {cumulative_return:+.2%}")

                    continue  # Skip to next date

            logger.info(f"✓ REBALANCING - Drift {total_drift:.2%} exceeds threshold" if len(current_portfolio) > 0 else "✓ INITIAL PORTFOLIO CREATION")

            # Generate trades
            trades = self._generate_trades(current_portfolio, new_portfolio, date)
            logger.info(f"Trades: {len(trades)} ({(trades['action']=='BUY').sum()} BUY, {(trades['action']=='SELL').sum()} SELL)")

            # Store results
            portfolio_with_date = new_portfolio.copy()
            portfolio_with_date.insert(0, "date", date)
            self.portfolio_history.append(portfolio_with_date)

            if len(trades) > 0:
                self.trades_history.append(trades)

            # Calculate performance metrics
            portfolio_value = new_portfolio["value"].sum()

            # Track previous portfolio value for period return calculation
            if i == 0:
                period_return = 0.0
                cumulative_return = 0.0
            else:
                # Get previous portfolio value
                prev_data = self.performance_history[-1]
                prev_portfolio_value = prev_data["portfolio_value"]

                # Calculate period return based on actual performance (excluding capital additions)
                # Period return = (ending value - beginning value - capital added) / beginning value
                period_return = (portfolio_value - prev_portfolio_value - capital_added_this_period) / prev_portfolio_value if prev_portfolio_value > 0 else 0

                # Cumulative return is profit/loss relative to total contributions
                cumulative_return = (portfolio_value - total_contributions) / total_contributions

            self.performance_history.append(
                {
                    "date": date,
                    "portfolio_value": portfolio_value,
                    "contributions": total_contributions,
                    "period_return": period_return,
                    "cumulative_return": cumulative_return,
                    "num_positions": len(new_portfolio),
                }
            )

            logger.info(f"Portfolio value: ${portfolio_value:,.2f}")
            logger.info(f"Period return: {period_return:+.2%}")
            logger.info(f"Cumulative return: {cumulative_return:+.2%}")

            # Update current portfolio
            current_portfolio = new_portfolio

        logger.info("=" * 80)
        logger.info("BACKTEST COMPLETED")
        logger.info("=" * 80)

        return True

    def generate_results(self) -> dict:
        """Calculate final performance metrics"""
        if len(self.performance_history) == 0:
            return {}

        perf_df = pd.DataFrame(self.performance_history)
        perf_df.set_index("date", inplace=True)

        # Calculate returns
        returns = perf_df["period_return"].values
        final_value = perf_df["portfolio_value"].iloc[-1]
        initial_value = self.initial_capital
        total_contributions = perf_df["contributions"].iloc[-1]

        # Time-weighted return (CAGR)
        days = (perf_df.index[-1] - perf_df.index[0]).days
        years = days / 365.25
        total_return = (final_value - total_contributions) / total_contributions
        cagr = (1 + total_return) ** (1 / years) - 1 if years > 0 else 0

        # Volatility and Sharpe
        volatility = np.std(returns) * np.sqrt(52)  # Weekly to annual
        mean_return = np.mean(returns) * 52  # Weekly to annual
        sharpe = mean_return / volatility if volatility > 0 else 0

        # Max drawdown
        cumulative_values = perf_df["portfolio_value"].values
        peak = np.maximum.accumulate(cumulative_values)
        drawdown = (cumulative_values - peak) / peak
        max_drawdown = np.min(drawdown)

        # Count rebalances
        total_rebalances = len(self.trades_history)

        results = {
            "start_date": perf_df.index[0],
            "end_date": perf_df.index[-1],
            "initial_capital": initial_value,
            "total_contributions": total_contributions,
            "final_value": final_value,
            "total_return": total_return,
            "cagr": cagr,
            "volatility": volatility,
            "sharpe_ratio": sharpe,
            "max_drawdown": max_drawdown,
            "total_rebalances": total_rebalances,
            "avg_positions": perf_df["num_positions"].mean(),
        }

        return results

    def save_to_excel(self, filename: str):
        """Save backtest results to Excel"""
        logger.info(f"Saving results to {filename}...")

        wb = Workbook()
        wb.remove(wb.active)

        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        results = self.generate_results()

        summary_data = [
            ["Metric", "Value"],
            ["Start Date", results["start_date"].strftime("%Y-%m-%d")],
            ["End Date", results["end_date"].strftime("%Y-%m-%d")],
            ["Initial Capital", f"${results['initial_capital']:,.2f}"],
            ["Total Contributions", f"${results['total_contributions']:,.2f}"],
            ["Final Value", f"${results['final_value']:,.2f}"],
            ["Total Return", f"{results['total_return']:.2%}"],
            ["CAGR", f"{results['cagr']:.2%}"],
            ["Volatility", f"{results['volatility']:.2%}"],
            ["Sharpe Ratio", f"{results['sharpe_ratio']:.2f}"],
            ["Max Drawdown", f"{results['max_drawdown']:.2%}"],
            ["Total Rebalances", results['total_rebalances']],
            ["Avg Positions", f"{results['avg_positions']:.1f}"],
        ]

        for row in summary_data:
            ws_summary.append(row)

        # Style summary sheet
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 20

        # Performance sheet
        ws_perf = wb.create_sheet("Performance")
        perf_df = pd.DataFrame(self.performance_history)

        # Remove timezone from date column
        if 'date' in perf_df.columns:
            perf_df['date'] = perf_df['date'].dt.tz_localize(None)

        for r_idx, row in enumerate(dataframe_to_rows(perf_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_perf.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        # Trades sheet
        if len(self.trades_history) > 0:
            ws_trades = wb.create_sheet("Trades")
            all_trades = pd.concat(self.trades_history, ignore_index=True)

            # Remove timezone from date column
            if 'date' in all_trades.columns:
                all_trades['date'] = all_trades['date'].dt.tz_localize(None)

            for r_idx, row in enumerate(dataframe_to_rows(all_trades, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_trades.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font

        # Portfolio history sheet
        if len(self.portfolio_history) > 0:
            ws_portfolio = wb.create_sheet("Portfolio_History")
            all_portfolios = pd.concat(self.portfolio_history, ignore_index=True)

            # Remove timezone from date column
            if 'date' in all_portfolios.columns:
                all_portfolios['date'] = all_portfolios['date'].dt.tz_localize(None)

            for r_idx, row in enumerate(
                dataframe_to_rows(all_portfolios, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_portfolio.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font

        wb.save(filename)
        logger.info(f"Results saved to {filename}")


def main():
    parser = argparse.ArgumentParser(description="Run paper trading backtest")
    parser.add_argument(
        "--start",
        type=str,
        required=True,
        help="Start date (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--end",
        type=str,
        default=datetime.now().strftime("%Y-%m-%d"),
        help="End date (YYYY-MM-DD, default: today)",
    )
    parser.add_argument(
        "--capital",
        type=float,
        default=100000,
        help="Initial capital (default: 100000)",
    )
    parser.add_argument(
        "--monthly-add",
        type=float,
        default=0,
        help="Monthly capital addition in last week (default: 0)",
    )
    parser.add_argument(
        "--positions",
        type=int,
        default=20,
        help="Number of positions (default: 20)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output Excel file (default: results/paper_trading_backtest_YYYYMMDD.xlsx)",
    )
    parser.add_argument(
        "--factor-ema-alpha",
        type=float,
        default=0.3,
        help="Factor score EMA smoothing alpha (default: 0.3, lower = more smoothing)",
    )
    parser.add_argument(
        "--drift-threshold",
        type=float,
        default=0.15,
        help="Portfolio drift threshold for rebalancing (default: 0.15 = 15%%)",
    )
    parser.add_argument(
        "--no-real-expense-ratios",
        action="store_true",
        help="Use synthetic expense ratios instead of fetching real ones",
    )

    args = parser.parse_args()

    # Create output filename
    if args.output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = RESULTS_DIR / f"paper_trading_backtest_{timestamp}.xlsx"
    else:
        output_file = Path(args.output)

    # Run backtest
    backtest = PaperTradingBacktest(
        start_date=args.start,
        end_date=args.end,
        initial_capital=args.capital,
        monthly_addition=args.monthly_add,
        num_positions=args.positions,
        factor_ema_alpha=args.factor_ema_alpha,
        drift_threshold=args.drift_threshold,
        use_real_expense_ratios=not args.no_real_expense_ratios,
    )

    success = backtest.run()

    if success:
        # Save results
        backtest.save_to_excel(str(output_file))

        # Print summary
        results = backtest.generate_results()
        print("\n" + "=" * 80)
        print("BACKTEST SUMMARY")
        print("=" * 80)
        print(f"Period: {results['start_date'].date()} to {results['end_date'].date()}")
        print(f"Initial Capital: ${results['initial_capital']:,.2f}")
        print(f"Total Contributions: ${results['total_contributions']:,.2f}")
        print(f"Final Value: ${results['final_value']:,.2f}")
        print(f"Total Return: {results['total_return']:.2%}")
        print(f"CAGR: {results['cagr']:.2%}")
        print(f"Volatility: {results['volatility']:.2%}")
        print(f"Sharpe Ratio: {results['sharpe_ratio']:.2f}")
        print(f"Max Drawdown: {results['max_drawdown']:.2%}")
        print(f"Total Rebalances: {results['total_rebalances']}")
        print(f"Avg Positions: {results['avg_positions']:.1f}")
        print("=" * 80)
        print(f"\nDetailed results saved to: {output_file}")

        sys.exit(0)
    else:
        logger.error("Backtest failed")
        sys.exit(1)


if __name__ == "__main__":
    main()
